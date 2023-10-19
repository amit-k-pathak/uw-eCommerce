from flask import *
import sqlite3, hashlib, os
from werkzeug.utils import secure_filename
from datetime import datetime
from itsdangerous import URLSafeTimedSerializer
from flask_mail import Mail, Message
# import xlsxwriter module
from openpyxl import load_workbook
import re

app = Flask(__name__)
app.secret_key = 'random string'
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = set(['jpeg', 'jpg', 'png', 'gif'])
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

app.config["SECURITY_PASSWORD_SALT"] = "email-confirm-key"
app.config["SECRET_KEY"] = "SECRET_KEY"
app.config["MAIL_DEFAULT_SENDER"]="noreply@flask.com"
app.config["MAIL_SERVER"] = "smtp.zoho.in"
app.config["MAIL_PORT"] = 465
app.config["MAIL_USE_TLS"] = False
app.config["MAIL_USE_SSL"] = True
app.config["MAIL_DEBUG"] = False
app.config["MAIL_USERNAME"] = "amit.pathak123@zohomail.in"
app.config["MAIL_PASSWORD"] = "744accounts.zoho.in"

def getLoginDetails():
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        if 'email' not in session:
            loggedIn = False
            firstName = ''
            noOfItems = 0
        else:
            loggedIn = True
            # cur.execute("SELECT userId, firstName FROM users WHERE email = ?", (session['email'],))
            cur.execute("SELECT userId, CompanyName FROM customers WHERE email = ?", (session['email'],))
            userId, firstName = cur.fetchone()
            # cur.execute("SELECT count(DISTINCT (productId)) FROM kart WHERE userId = ?", (userId, ))
            cur.execute("SELECT count(DISTINCT (productId)) FROM kart WHERE userId = ? and OrderID=?", (userId, ""))
            noOfItems = cur.fetchone()[0]
    conn.close()
    return (loggedIn, firstName, noOfItems)

@app.route("/")

def root():
    loggedIn, firstName, noOfItems = getLoginDetails()
    with sqlite3.connect('database.db') as conn:
            cur = conn.cursor()
            # cur.execute('SELECT productId, name, price, description, image, stock FROM products')
            if loggedIn:
                cName = get_customer_catagory(session['email'])
            else:
                cName = ""
            cur.execute('SELECT productId, name, pricepercigar, priceperbox, image FROM allProducts where catagory=?', (cName,))
            itemData = cur.fetchall()
            cur.execute('SELECT categoryId, name FROM categories')
            categoryData = cur.fetchall()
    itemData = parse(itemData)
    return render_template('home.html', itemData=itemData, loggedIn=loggedIn, firstName=firstName, noOfItems=noOfItems,
                           categoryData=categoryData)

@app.route("/add")
def adminAddProduct():
    loggedIn, firstName, noOfItems = getLoginDetails()
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute("SELECT categoryId, name FROM categories")
        categories = cur.fetchall()
    conn.close()
    # return render_template('add.html', categories=categories)
    return render_template('addProduct.html', categories=categories, loggedIn=loggedIn, firstName=firstName,noOfItems=noOfItems)

# Admin side listing of all available products
@app.route("/productList")
def adminListProducts():
    loggedIn, firstName, noOfItems = getLoginDetails()
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute('SELECT productId, name, size, weight, presentation, count, pricepercigar, priceperbox, countryoforigin, image FROM allProducts')
        productData = cur.fetchall()
    conn.close()
    return render_template("listProducts.html", data=productData, loggedIn=loggedIn, firstName=firstName,noOfItems=noOfItems)

# Admin side listing of all orders
@app.route("/orderList")
def adminOrderList():
    loggedIn, firstName, noOfItems = getLoginDetails()
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute('SELECT OrderId, OrderDate, UserEmail, TotalPrice, OrderStatus FROM Orders order by OrderDate desc')
        orderData = cur.fetchall()
    conn.close()
    return render_template("adminOrders.html", data=orderData, loggedIn=loggedIn, firstName=firstName,noOfItems=noOfItems)

# Admin side Updation of Product Data
@app.route("/update", methods=['GET', 'POST'])
def update():
    # productid = request.args.get('productId')
    productid = request.form['id']
    if request.method == 'POST':
        pName = request.form["name"]
        pDescrition = request.form["description"]
        pPrice= request.form["price"]
        pStock= request.form["stock"]

        # Uploading image procedure
        image = request.files['image']
        filename = ""
        if image and allowed_file(image.filename):
            filename = secure_filename(image.filename)
            image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        imagename = filename

        with sqlite3.connect('database.db') as con:
            try:
                cur = con.cursor()

                if len(imagename.strip()) > 0:
                    cur.execute(
                        'UPDATE products SET name = ?, price = ?, description = ?, stock = ?, image=? WHERE productId = ?',
                        (pName, pPrice, pDescrition, pStock, imagename, productid))
                else:
                    cur.execute(
                        'UPDATE products SET name = ?, price = ?, description = ?, stock = ? WHERE productId = ?',
                        (pName, pPrice, pDescrition, pStock, productid))

                flash("Product Edits Saved Successfully.")
                con.commit()
            except:
                con.rollback()
                flash("Error occured!")
        con.close()
    return redirect(url_for('adminListProducts'))

# Admin side Updation of Order Data
@app.route("/updateOrder", methods=['GET', 'POST'])
def updateOrder():
    # productid = request.args.get('productId')
    orderid = request.form['id']
    if request.method == 'POST':

        orderStatus= request.form["status"]

        with sqlite3.connect('database.db') as con:
            try:
                cur = con.cursor()
                cur.execute(
                    'UPDATE Orders SET OrderStatus = ? WHERE orderId = ?',
                    (orderStatus, orderid))

                flash("Order Edited Successfully.")
                con.commit()
            except:
                con.rollback()
                flash("Error occured!")
        con.close()
    return redirect(url_for('adminOrderList'))

# Admin side Updation of Customer Data
@app.route("/updateCustomer", methods=['GET', 'POST'])
def updateCustomer():
    # productid = request.args.get('productId')
    userid = request.form['id']
    if request.method == 'POST':

        cat = request.form["catagory"]

        with sqlite3.connect('database.db') as con:
            try:
                cur = con.cursor()
                cur.execute(
                    'UPDATE customers SET catagory = ? WHERE userId = ?',
                    (cat, userid))

                flash("Customer Edited Successfully.")
                con.commit()
            except:
                con.rollback()
                flash("Error occured!")
        con.close()
    return redirect(url_for('adminListCustomers'))








@app.route("/checkout")
def checkoutPlaceOrder():
    if 'email' not in session:
        return redirect(url_for('root'))
    loggedIn, firstName, noOfItems = getLoginDetails()
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        # cur.execute(
        #     "SELECT userId, email, firstName, lastName, address1, address2, zipcode, city, state, country, phone FROM users WHERE email = ?",
        #     (session['email'],))
        cur.execute(
            "SELECT userId, email, CompanyName, BillingAddress1, BillingAddress2, BillingZipCode, BillingCity, BillingState, BillingCountry, BillingPhoneNumber FROM customers WHERE email = ?",
            (session['email'],))


        profileData = cur.fetchone()
    conn.close()

    if noOfItems > 0:
        # create order
        orderId = 0
        dTime = datetime.utcnow()
        totalPriceTpl = getTotalPrice()
        with sqlite3.connect('database.db') as conn:
            try:
                cur = conn.cursor()
                cur.execute(
                    '''INSERT INTO Orders (OrderDate, UserEmail, TotalPrice, OrderStatus) VALUES (?,?,?,?)''',
                    (dTime, session['email'],totalPriceTpl[0], 'PENDING'))
                orderId = cur.lastrowid
                conn.commit()
                msg = "Order created successfully."
            except:
                msg = "error occured!"
                conn.rollback()
        conn.close()

        # update kart
        with sqlite3.connect('database.db') as con:
            try:
                cur = con.cursor()
                cur.execute(
                    'UPDATE kart SET OrderID = ? WHERE userid = ? and OrderID = ?',
                    (orderId, profileData[0], ""))
                con.commit()
                msg = "Order placed successfully."
            except:
                con.rollback()
                msg = "Error occured!"
        con.close()
        # Remove Items from kart
        noOfItems = 0

        #For each product in order, update stock in products table
        # Update Product Stock
        with sqlite3.connect('database.db') as con:
            try:
                cur = con.cursor()
                cur.execute(
                    'SELECT productId from kart WHERE OrderID = ? and userid = ?',
                    (orderId, profileData[0]))
                rows = cur.fetchall()
                for r in rows:
                    cur = con.cursor()
                    cur.execute(
                        'UPDATE products set stock = stock-1 WHERE productId=?',
                        (r[0],))

                con.commit()
            except:
                con.rollback()
                msg = "Error occured!"
        con.close()

        # Send an e-mail notifying warehouse
        html = render_template("order_placed.html")
        subject = "Order Placed"
        send_email("s.lampert@outlook.com", subject, html)

        msg = "Your order is placed and our warehouse has been notified. Our warehouse will fulfill your order as soon as possible."

    else:
        msg = "Your shopping cart is empty! Please add some products and checkout."

    return render_template("checkoutPlaceOrder.html", profileData=profileData, loggedIn=loggedIn, firstName=firstName,
                           noOfItems=noOfItems, msg=msg)
    # with sqlite3.connect('database.db') as conn:
    #     cur = conn.cursor()
    #     cur.execute("SELECT categoryId, name FROM categories")
    #     categories = cur.fetchall()
    # conn.close()
    # return render_template('checkoutPlaceOrder.html', categories=categories)

#Admin Side Addition of a Product
@app.route("/addItem", methods=["GET", "POST"])
def addItem():
    if request.method == "POST":

        name = request.form['productName']
        size = request.form['size']
        weight = request.form['weight']
        presentation = request.form['presentation']
        count = request.form['count']
        pricepercigar = request.form['pricepercigar']
        priceperbox = request.form['priceperbox']
        countryoforigin = request.form['countryoforigin']
        category = request.form['category']

        # Uploading image procedure
        image = request.files['image']
        if image and allowed_file(image.filename):
            filename = secure_filename(image.filename)
            image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        imagename = filename

        with sqlite3.connect('database.db') as conn:
            try:
                cur = conn.cursor()
                # cur.execute(
                #     '''INSERT INTO products (name, price, description, image, stock, categoryId) VALUES (?, ?, ?, ?, ?, ?)''',
                #     (name, price, description, imagename, stock, categoryId))

                cur.execute(
                    '''INSERT INTO allProducts (name, "size", weight, presentation, "count", pricepercigar, priceperbox, countryoforigin, catagory)  VALUES (?, ?, ?, ?, ?,?, ?, ?, ?)''',
                    (name, size, weight, presentation, count, pricepercigar, priceperbox, countryoforigin, category))


                conn.commit()
                msg = "Product added successfully."
            except:
                msg = "An error occured during processing!"
                conn.rollback()
        conn.close()
        print(msg)
        return redirect(url_for('adminListProducts'))

@app.route("/remove")
def remove():
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute('SELECT productId, name, price, description, image, stock FROM products')
        data = cur.fetchall()
    conn.close()
    return render_template('remove.html', data=data)

#Admin Side deletion of a Product
@app.route("/delete/<id>", methods=['GET', 'POST'])
def delete(id):
    with sqlite3.connect('database.db') as conn:
        try:
            cur = conn.cursor()
            cur.execute('DELETE FROM products WHERE productID = ?', (id,))
            conn.commit()
            flash("Deleted successsfully.")
            conn.commit()
        except:
            conn.rollback()
            flash("Error occured!")
    conn.close()

    return redirect(url_for('adminListProducts'))

#Admin Side deletion of an Order
@app.route("/deleteOrder/<id>", methods=['GET', 'POST'])
def deleteOrder(id):
    with sqlite3.connect('database.db') as conn:
        try:
            cur = conn.cursor()
            cur.execute('DELETE FROM Orders WHERE OrderId = ?', (id,))
            conn.commit()
            flash("Deleted successsfully.")
            conn.commit()
        except:
            conn.rollback()
            flash("Error occured!")
    conn.close()

    return redirect(url_for('adminOrderList'))

#Admin Side deletion of an Customer
@app.route("/deleteCustomer/<id>", methods=['GET', 'POST'])
def deleteCustomer(id):
    with sqlite3.connect('database.db') as conn:
        try:
            cur = conn.cursor()
            cur.execute('DELETE FROM customers WHERE userId = ?', (id,))
            conn.commit()
            flash("Customer Deleted successsfully.")
            conn.commit()
        except:
            conn.rollback()
            flash("Error occured!")
    conn.close()

    return redirect(url_for('adminListCustomers'))

@app.route("/displayCategory")
def displayCategory():
    loggedIn, firstName, noOfItems = getLoginDetails()
    categoryId = request.args.get("categoryId")
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT products.productId, products.name, products.price, products.image, categories.name FROM products, categories WHERE products.categoryId = categories.categoryId AND categories.categoryId = ?",
            (categoryId,))
        data = cur.fetchall()
    conn.close()
    categoryName = ""
    if len(data) > 0:
         categoryName = data[0][4]
         data = parse(data)
    else:
        with sqlite3.connect('database.db') as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM categories WHERE categoryId = ?",
                (categoryId,))
            categoryName = cur.fetchall()
        conn.close()
        categoryName= categoryName[0][0]

    return render_template('displayCategory.html', data=data, loggedIn=loggedIn, firstName=firstName,
                           noOfItems=noOfItems, categoryName=categoryName)

@app.route("/account/profile")
def profileHome():
    if 'email' not in session:
        return redirect(url_for('root'))
    loggedIn, firstName, noOfItems = getLoginDetails()
    return render_template("profileHome.html", loggedIn=loggedIn, firstName=firstName, noOfItems=noOfItems)

@app.route("/account/profile/edit")
def editProfile():
    if 'email' not in session:
        return redirect(url_for('root'))
    loggedIn, firstName, noOfItems = getLoginDetails()
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT userId, email, CompanyName, BillingAddress1, BillingAddress2, BillingZipCode, BillingCity, BillingState, BillingCountry, BillingPhoneNumber, ShippingAddress1,ShippingAddress2, ShippingZipCode, ShippingCity, ShippingState, ShippingCountry, ShippingPhoneNumber FROM customers WHERE email = ?",
            (session['email'],))
        profileData = cur.fetchone()
    conn.close()
    return render_template("editProfile.html", profileData=profileData, loggedIn=loggedIn, firstName=firstName,
                           noOfItems=noOfItems)

# List all orders made by loggedIn User
@app.route("/account/orders")
def showOrders():
    if 'email' not in session:
        return redirect(url_for('root'))
    loggedIn, firstName, noOfItems = getLoginDetails()
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute("SELECT OrderID, OrderDate, UserEmail, TotalPrice, OrderStatus FROM Orders WHERE UserEmail = ? order by OrderDate desc", (session['email'],))
        ordersData = cur.fetchall()
        # //conn.close()

        if len(ordersData) == 0:
            msg = "No orders exist in records!"
        else:
            msg = "Found orders in records!"

    return render_template("listOrders.html", profileData=ordersData, loggedIn=loggedIn, firstName=firstName,
                           noOfItems=noOfItems, msg=msg)

# List all customers present in DB
@app.route("/listCustomers")
def adminListCustomers():
    if 'email' not in session:
        return redirect(url_for('root'))
    loggedIn, firstName, noOfItems = getLoginDetails()

    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        # cur.execute(
        #     "SELECT userId, email, firstName, address1, zipcode, city, country, phone, IsAdmin FROM users")
        cur.execute(
            "SELECT userId, email, CompanyName, BillingAddress1, BillingZipCode, BillingCity, BillingCountry, BillingPhoneNumber, IsAdmin, catagory FROM customers")

        profileData = cur.fetchall()
    conn.close()

    return render_template("listCustomers.html", profileData=profileData, loggedIn=loggedIn, firstName=firstName,
                           noOfItems=noOfItems)

@app.route("/account/profile/view")
def viewProfile():
    if 'email' not in session:
        return redirect(url_for('root'))
    loggedIn, firstName, noOfItems = getLoginDetails()
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT userId, email, CompanyName, BillingAddress1, BillingAddress2, BillingZipCode, BillingCity, BillingState, BillingCountry, BillingPhoneNumber, ShippingAddress1, ShippingAddress2, ShippingZipCode, ShippingCity, ShippingState, ShippingCountry, ShippingPhoneNumber FROM customers WHERE email = ?",
            (session['email'],))
        profileData = cur.fetchone()
    conn.close()
    return render_template("viewProfile.html", profileData=profileData, loggedIn=loggedIn, firstName=firstName,
                           noOfItems=noOfItems)

@app.route("/account/profile/changePassword", methods=["GET", "POST"])
def changePassword():
    if 'email' not in session:
        return redirect(url_for('loginForm'))
    if request.method == "POST":
        oldPassword = request.form['oldpassword']
        oldPassword = hashlib.md5(oldPassword.encode()).hexdigest()
        newPassword = request.form['newpassword']
        newPassword = hashlib.md5(newPassword.encode()).hexdigest()
        with sqlite3.connect('database.db') as conn:
            cur = conn.cursor()
            # cur.execute("SELECT userId, password FROM users WHERE email = ?", (session['email'],))
            cur.execute("SELECT userId, password FROM customers WHERE email = ?", (session['email'],))

            userId, password = cur.fetchone()
            if (password == oldPassword):
                try:
                    # cur.execute("UPDATE users SET password = ? WHERE userId = ?", (newPassword, userId))
                    cur.execute("UPDATE customers SET password = ? WHERE userId = ?", (newPassword, userId))
                    conn.commit()
                    msg = "Password Changed successfully"
                except:
                    conn.rollback()
                    msg = "Couldn't Change Password!"
                return render_template("changePassword.html", msg=msg)
            else:
                msg = "Incorrect old password."
        conn.close()
        return render_template("changePassword.html", msg=msg)
    else:
        loggedIn, firstName, noOfItems = getLoginDetails()
        return render_template("changePassword.html", loggedIn=loggedIn, firstName=firstName,noOfItems=noOfItems)
        # return render_template("changePassword.html")

@app.route("/updateProfile", methods=["GET", "POST"])
def updateProfile():
    if request.method == 'POST':
        email = request.form['email']
        companyName = request.form['companyName']
        baddress1 = request.form['billingaddress1']
        baddress2 = request.form['billingaddress2']
        bzipcode = request.form['billingzipcode']
        bcity = request.form['billingcity']
        bstate = request.form['billingstate']
        bcountry = request.form['billingcountry']
        bphone = request.form['billingphone']

        saddress1 = request.form['shippingaddress1']
        saddress2 = request.form['shippingaddress2']
        szipcode = request.form['shippingzipcode']
        scity = request.form['shippingcity']
        sstate = request.form['shippingstate']
        scountry = request.form['shippingcountry']
        sphone = request.form['shippingphone']

        with sqlite3.connect('database.db') as con:
            try:
                cur = con.cursor()
                # cur.execute(
                #     'UPDATE users SET firstName = ?, lastName = ?, address1 = ?, address2 = ?, zipcode = ?, city = ?, state = ?, country = ?, phone = ? WHERE email = ?',
                #     (firstName, lastName, address1, address2, zipcode, city, state, country, phone, email))
                cur.execute(
                    'UPDATE customers SET CompanyName = ?, BillingAddress1 = ?, BillingAddress2 = ?, BillingZipCode = ?, BillingCity = ?, BillingState = ?, BillingCountry = ?, BillingPhoneNumber = ?, ShippingAddress1 = ?, ShippingAddress2 = ?, ShippingZipCode = ?, ShippingCity = ?, ShippingState = ?, ShippingCountry = ?, ShippingPhoneNumber = ? WHERE email = ?',
                    (companyName, baddress1, baddress2, bzipcode, bcity, bstate, bcountry, bphone,saddress1, saddress2, szipcode, scity, sstate, scountry, sphone, email))

                con.commit()
                msg = "Profile Edits Saved Successfully."
            except:
                con.rollback()
                msg = "Error occured!"
        con.close()

        if 'email' not in session:
            return redirect(url_for('root'))
        loggedIn, firstName, noOfItems = getLoginDetails()

        with sqlite3.connect('database.db') as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT userId, email, CompanyName, BillingAddress1, BillingAddress2, BillingZipCode, BillingCity, BillingState, BillingCountry, BillingPhoneNumber FROM customers WHERE email = ?",
                (session['email'],))
            profileData = cur.fetchone()
        conn.close()
        return render_template("editProfile.html", profileData=profileData, loggedIn=loggedIn, firstName=firstName,
                               noOfItems=noOfItems, msg=msg)

        # return redirect(url_for('editProfile'))
        # return render_template('editProfile.html', msg=msg)

@app.route("/loginForm")
def loginForm():
    if 'email' in session:
        adminFlag = is_admin(session['email'])
        if adminFlag is True:
            return redirect(url_for('adminListProducts'))
        else:
            return redirect(url_for('root'))
    else:
        return render_template('login.html', error='')

@app.route("/login", methods=['POST', 'GET'])
def login():
    loggedIn, firstName, noOfItems = getLoginDetails()

    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        if is_valid(email, password):
            if(is_admin(email)):
                 session['email'] = email
                 # return render_template("listProducts.html", profileData=profileData, loggedIn=loggedIn,firstName=firstName,noOfItems=noOfItems)
                 return redirect(url_for('adminListProducts'))
            else:
                 session['email'] = email
                 return redirect(url_for('root'))
        else:
            error = 'Invalid UserId / Password'
            return render_template('login.html', error=error)

@app.route("/productDescription")
def productDescription():
    loggedIn, firstName, noOfItems = getLoginDetails()
    productId = request.args.get('productId')
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute('SELECT productId, name, priceperbox, size, image, count, weight,SKU, pricepercigar,countryoforigin, presentation FROM allProducts WHERE productId = ?',
                    (productId,))
        productData = cur.fetchone()
    conn.close()
    return render_template("productDescription.html", data=productData, loggedIn=loggedIn, firstName=firstName,
                           noOfItems=noOfItems)

@app.route("/addToCart",methods=['POST', 'GET'])
def addToCart():
    if 'email' not in session:
        return redirect(url_for('loginForm'))
    else:


        productId = int(request.args.get('productId'))

        with sqlite3.connect('database.db') as conn:
            cur = conn.cursor()
            cur.execute("SELECT userId FROM customers WHERE email = ?", (session['email'],))
            userId = cur.fetchone()[0]

            # cur.execute("SELECT stock FROM products WHERE productId = ?", (productId,))
            # stock = cur.fetchone()[0]

            cur.execute("SELECT count(productId) FROM kart WHERE userId = ? and productId = ? and OrderID = ?", (userId, productId,""))
            inCart = cur.fetchone()[0]

            # stock = stock - inCart
            # inStock = True if stock > 0 else False

            # if inStock is True:
            try:
                    # cur.execute("INSERT INTO kart (userId, productId) VALUES (?, ?)", (userId, productId))
                    cur.execute("INSERT INTO kart (userId, productId, OrderID) VALUES (?, ?, ?)", (userId, productId, ""))
                    # cur.execute("INSERT INTO ShoppingCart (userId, productId, quantity, OrderID) VALUES (?, ?, ?, ?)",(userId, productId, ""))
                    conn.commit()
                    msg = "Added successfully"
            except:
                    conn.rollback()
                    msg = "Error occured!"
            # else:
            #     msg = "Product is Out of Stock!"
        conn.close()
        return redirect(url_for('cart'))
# Under Modification on 05/09/2023
def getTotalPrice():
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute("SELECT userId FROM customers WHERE email = ?", (session['email'],))
        userId = cur.fetchone()[0]
        # cur.execute("SELECT products.productId, products.name, products.price, products.image FROM products, kart WHERE products.productId = kart.productId AND kart.userId = ?", (userId, ))
        # cur.execute(
        #     "SELECT products.productId, products.name, products.price, products.image, count(products.name) as Qty FROM products, kart WHERE products.productId = kart.productId AND kart.userId = ? AND kart.OrderID = ? group by products.name, products.productId",
        #     (userId, ""))
        cur.execute(
            "SELECT allProducts.productId, allProducts.name, allProducts.priceperbox, allProducts.image, count(allProducts.name) as Qty FROM allProducts, kart WHERE allProducts.productId = kart.productId AND kart.userId = ? AND kart.OrderID = ? group by allProducts.name, allProducts.productId",
            (userId, ""))
    products = cur.fetchall()

    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute("SELECT catagory FROM customers WHERE userId = ?", (userId,))
        catagoryId = cur.fetchone()[0]


    totalPrice = 0
    num = 0.0
    for row in products:
        num1 = re.findall("\d+\.?\d*", row[2])
        # num2 = re.findall("\d+\.?\d*", row[4])
        # totalPrice += round(float(num1[0]), 2) * int(row[4])
        totalPrice += round(float(num1[0]), 2) * int(row[4])
    currency = get_currency_by_catagory(catagoryId)
    totalPrice = currency + " " + str(round(totalPrice, 2))

    return (totalPrice, products)

def get_currency_by_catagory(c):
    if c == "FACTORY":
        return "$"
    elif c == "MY&MI EURO":
        return "€"
    elif c=="MY&MI USD":
        return  "$"
    else:
        return "CHF"

#
# def get_currency_by_catagory(c):
#     if c == "1":
#         return "$"
#     elif c == "2":
#         return "€"
#     elif c=="3":
#         return  "$"
#     else:
#         return "CHF"


def get_catagory_by_id(c):
    if c == "1":
        return "FACTORY"
    elif c == "MY&MI EURO":
        return "€"
    elif c=="3":
        return  "MY&MI USD"
    else:
        return "EXW SWITZERLAND CHF"


@app.route("/cart")
def cart():
    if 'email' not in session:
        return redirect(url_for('loginForm'))
    loggedIn, firstName, noOfItems = getLoginDetails()
    email = session['email']
    # with sqlite3.connect('database.db') as conn:
    #     cur = conn.cursor()
    #     cur.execute("SELECT userId FROM users WHERE email = ?", (email,))
    #     userId = cur.fetchone()[0]
    #     # cur.execute("SELECT products.productId, products.name, products.price, products.image FROM products, kart WHERE products.productId = kart.productId AND kart.userId = ?", (userId, ))
    #     cur.execute(
    #         "SELECT products.productId, products.name, products.price, products.image, count(products.name) as Qty FROM products, kart WHERE products.productId = kart.productId AND kart.userId = ? AND kart.OrderID = ? group by products.name",
    #         (userId, ""))
    #
    #     products = cur.fetchall()
    # totalPrice = 0
    # for row in products:
    #     totalPrice += (int(row[2]) * int(row[4]))
    if request.method == 'GET':
        qty = request.args.get('quantity')

        if qty == None:
            qty = 1

    totalPrice, products = getTotalPrice()
    return render_template("cart.html", products=products, totalPrice=totalPrice, loggedIn=loggedIn,
                           firstName=firstName, noOfItems=noOfItems, qty=qty)


@app.route("/removeFromCart")
def removeFromCart():
    if 'email' not in session:
        return redirect(url_for('loginForm'))
    email = session['email']
    productId = int(request.args.get('productId'))
    with sqlite3.connect('database.db') as conn:
        cur = conn.cursor()
        cur.execute("SELECT userId FROM customers WHERE email = ?", (email,))
        userId = cur.fetchone()[0]
        try:
            cur.execute("DELETE FROM kart WHERE userId = ? AND productId = ?", (userId, productId))
            conn.commit()
            msg = "removed successfully"
        except:
            conn.rollback()
            msg = "error occured"
    conn.close()
    return redirect(url_for('cart'))

@app.route("/logout")
def logout():
    session.pop('email', None)
    return redirect(url_for('root'))
def is_valid(email, password):
    con = sqlite3.connect('database.db')
    cur = con.cursor()

    cur.execute('SELECT email, password, IsConfirmed FROM customers')
    data = cur.fetchall()
    for row in data:
        if row[0] == email and row[1] == hashlib.md5(password.encode()).hexdigest() and row[2] == 1:
            return True
    return False
def is_admin(email):
    con = sqlite3.connect('database.db')
    cur = con.cursor()
    cur.execute('SELECT IsAdmin FROM customers where email=?', (email,))
    data = cur.fetchone()
    if data[0] == 1:
            return True
    return False

def get_customer_catagory(email):
    con = sqlite3.connect('database.db')
    cur = con.cursor()
    cur.execute('SELECT catagory FROM customers where email=?', (email,))
    data = cur.fetchone()
    return data[0]

def write_excel_and_send_email(email):
    con = sqlite3.connect('database.db')
    cur = con.cursor()
    cur.execute('SELECT *  from customers c natural join invoicesAndDocuments iad WHERE c.email = ?', (email,))
    data = cur.fetchone()

    wb = load_workbook('C:/Users/ok/Desktop/UW/Website-Stefan/Kopie von New Customer Form.xlsx')

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['R7C2'] = data[0] # data[2]?
# Billing
    ws['R9C2'] = data[1]
    ws['R10C2'] = data[2]
    ws['R11C2'] = data[3]
    ws['R12C2'] = data[4]
    ws['R13C2'] = data[5]
    ws['R14C2'] = data[6]
    ws['R15C2'] = data[7]
    ws['R16C2'] = data[8]
    ws['R17C2'] = data[9]

# Shipping Address
    ws['R19C2'] = data[10]
    ws['R20C2'] = data[11]
    ws['R21C2'] = data[12]
    ws['R22C2'] = data[13]
    ws['R23C2'] = data[14]
    ws['R24C2'] = data[15]
    ws['R25C2'] = data[16]
    ws['R26C2'] = data[17]
    ws['R27C2'] = data[18]

    ws['R29C2'] = 'Company Name'
    ws['R29C9'] = 'Company Name'
    ws['R30C2'] = 'Company Name'
    ws['R30C9'] = 'Company Name'
    ws['R31C2'] = 'Company Name'
    ws['R31C9'] = 'Company Name'
# Required notes on invoice
    ws['R33C1'] = ''

    # Required Export Documents
    ws['R38C2'] = 'Company Name'
    ws['R38C9'] = 'Company Name'
    ws['R39C2'] = 'Company Name'
    ws['R39C9'] = 'Company Name'
    ws['R40C2'] = 'Company Name'
    ws['R40C9'] = 'Company Name'
    ws['R41C2'] = 'Company Name'
    ws['R41C9'] = 'Company Name'
    ws['R42C2'] = 'Company Name'
    ws['R42C9'] = 'Company Name'


    # Save the file
    wb.save("template_filled.xlsx")

    # Finally, close the Excel file
    # via the close() method.
    wb.close()

    send_email('amit180889@gmail.com', 'New Customer Registered!',None , "template_filled.xlsx")

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
@app.route("/confirm/<token>")
def confirm_email(token):
    email = confirm_token(token)

    with sqlite3.connect('database.db') as con:
            cur = con.cursor()
            cur.execute(
                'SELECT IsConfirmed from customers where email=?',
                (email,))
            isConfirmed=cur.fetchone()[0]

            if isConfirmed == 0:
                confirmed_on = datetime.utcnow()
                with sqlite3.connect('database.db') as con:
                    try:
                        cur = con.cursor()
                        cur.execute(
                            'UPDATE customers set IsConfirmed=1, ConfirmedOn=? where email=?',
                            (confirmed_on, email))
                        con.commit()
                    except:
                        con.rollback()
                con.close()
                #
                # write_excel_and_send_email(email)
                #
            else:
                msg = "Account already confirmed."
                return render_template('login.html', error=msg)

        # flash("You have confirmed your account. Thanks!", "success")
    # else:
    #     flash("The confirmation link is invalid or has expired.", "danger")
    return render_template('login.html', error='')
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

@app.route("/invite", methods=['GET','POST'])
def invite():
    if request.method == 'POST':
        email = request.form['recipientEmail']
        invt_msg = request.form['invitationMessage']
        try:
            token = generate_token(email)
            url = url_for("register", token=token, _external=True)
            html = render_template("invitation.html", url=url)
            subject = "Please register to website"
            send_email(email, subject, html)
            msg = "An invitation has been sent via email. Please check your e-mail ID for registraction."

            return render_template("inviteForm.html", msg=msg)
        except:
            msg = "Error occured!"
            return render_template("inviteForm.html", msg=msg)
    else:
        msg = ""
        return render_template("inviteForm.html", msg=msg)

# New Register Function
@app.route("/register", methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # Parse form data
        password = request.form['password']
        email = request.form['email']
        ##########################################################
        companyName = request.form['companyName']
        ##########################################################
        ## Get Billing Address
        address1 = request.form['addressLine1']
        address2 = request.form['addressLine2']
        zipcode = request.form['zipCode']
        city = request.form['city']
        state = request.form['state']
        country = request.form['country']
        contactPerson = request.form['contactPerson']
        phone = request.form['phone']
        billingEmail = request.form['billingEmailId']
        ##########################################################
        ## Get Shipping Address
        shippingaddress1 = request.form['shippingAddressLine1']
        shippingaddress2 = request.form['shippingAddressLine2']
        shippingzipcode = request.form['shippingZipCode']
        shippingcity = request.form['shippingCity']
        shippingstate = request.form['shippingState']
        shippingcountry = request.form['shippingCountry']
        shippingcontactperson = request.form['shippingContactPerson']
        shippingphone = request.form['shippingPhone']
        shippingemail = request.form['shippingEmailId']
        ##########################################################
        boxes = 1 if request.form['boxes'] == "yes" else 0
        toCigars = 1 if request.form['thousandsofCigars'] == "yes" else 0
        cigars = 1 if request.form['cigars'] == "yes" else 0
        kg = 1 if request.form['kilograms'] == "yes" else 0
        bundles = 1 if request.form['bundles'] == "yes" else 0
        other = 1 if request.form['other'] == "yes" else 0
        ##########################################################
        notes = request.form['notes']
        ##########################################################
        export = 1 if request.form['Export'] == "yes" else 0
        GSTP = 1 if request.form['GSTP'] == "yes" else 0
        Shipping = 1 if request.form['Shipping'] == "yes" else 0
        Stamps = 1 if request.form['Stamps'] == "yes" else 0
        Origin = 1 if request.form['Origin'] == "yes" else 0
        Phitosanitary = 1 if request.form['Phitosanitary'] == "yes" else 0
        EUR1 = 1 if request.form['EUR1'] == "yes" else 0
        License = 1 if request.form['License'] == "yes" else 0
        GSP = 1 if request.form['GSP'] == "yes" else 0
        other2 = 1 if request.form['other2'] == "yes" else 0
        ##########################################################

        con = sqlite3.connect('database.db')
        cur = con.cursor()
        cur.execute('SELECT email FROM customers where email=?', (email,))
        data = cur.fetchone()
        con.close()

        if data is None:
            with sqlite3.connect('database.db') as con:
                try:
                    cur = con.cursor()
                    # query has been modified here to replace First Name/Last Name
                    # cur.execute(
                    #     'INSERT INTO users (password, email, firstName, lastName, address1, address2, zipcode, city, state, country, phone, IsAdmin, IsConfirmed, ConfirmedOn) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                    #     (
                    #     hashlib.md5(password.encode()).hexdigest(), email, companyName, "", address1, address2, zipcode,
                    #     city, state, country, phone, 0, 0, ""))
                    cur.execute(
                        'INSERT INTO customers (password, email, BillingAddress1, BillingAddress2, BillingZipCode, BillingCity, BillingState, BillingCountry, BillingContactPerson,BillingPhoneNumber, BillingEmail, ShippingAddress1, ShippingAddress2, ShippingZipCode, ShippingCity, ShippingState, ShippingCountry, ShippingContactPerson,ShippingPhoneNumber, ShippingEmail, IsAdmin, IsConfirmed, ConfirmedOn, CompanyName) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                        (
                            hashlib.md5(password.encode()).hexdigest(), email, address1, address2,zipcode,city, state, country, contactPerson, phone, billingEmail, shippingaddress1, shippingaddress2, shippingzipcode, shippingcity, shippingstate, shippingcountry, shippingcontactperson, shippingphone, shippingemail, 0, 0, "",companyName))

                    # Retrieve newly inserted UserID
                    userId = cur.lastrowid
                    con.commit()
                    cur = con.cursor()

                    cur.execute(
                        'insert into invoicesAndDocuments (InvoiceUnitBoxes, InvoiceUnitThousandsofCigars, InvoiceUnitCigars, InvoiceUnitKilograms, InvoiceUnitBundles, InvoiceUnitOther, RequiredNotesOnInvoice, ExportInvoice, GSTPCertificate, ShippingList, TaxStamps, CertificateOfOrigin, PhitosanitaryCertificate, EUR1Cerificate, ImportLicense, GSPCertificate, Other, userId) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                        (
                            boxes, toCigars, cigars, kg, bundles, other, '', export, GSTP, Shipping, Stamps, Origin, Phitosanitary, EUR1, License, GSP, other2, userId))
                    con.commit()


                    token = generate_token(email)
                    confirm_url = url_for("confirm_email", token=token, _external=True)
                    html = render_template("confirm_email.html", confirm_url=confirm_url)
                    subject = "Please confirm your email"
                    send_email(email, subject, html)
                    msg = "A confirmation email has been sent via email."
                    # return render_template("register.html", error=msg)
                    return render_template("Reg-6.html", error=msg)
                except:
                    con.rollback()
                    msg = "Error occured during new customer registration!"
        else:
            con = sqlite3.connect('database.db')
            cur = con.cursor()
            cur.execute('SELECT IsConfirmed FROM customers where email=?', (email,))
            data = cur.fetchone()
            con.close()
            if data[0] == 0:
                msg = "It seems  You have not confirmed your account. Please check your inbox (and your spam folder) - you should have received an email with a confirmation link."
            else:
                msg = "Emaild already exist in our records, please select another emaildID!"
        con.close()
        # return render_template("register.html", error=msg)
        return render_template("Reg-6.html", error=msg)
    else:
        msg = ""
        return render_template("Reg-6.html", error=msg)
##############################################################################################
# @app.route("/register", methods=['GET', 'POST'])
# def register():
#     if request.method == 'POST':
#         # Parse form data
#         password = request.form['password']
#         email = request.form['email']
#         # firstName = request.form['firstName']
#         # lastName = request.form['lastName']
#         companyName = request.form['companyName']
#
#         ## Get Billing Address
#         address1 = request.form['address1']
#         address2 = request.form['address2']
#         zipcode = request.form['zipcode']
#         city = request.form['city']
#         state = request.form['state']
#         country = request.form['country']
#         contactPerson = request.form['contactPerson']
#         phone = request.form['phone']
#         ## Get Shipping Address
#         shippingaddress1 = request.form['shippingAddressLine1']
#         shippingaddress2 = request.form['shippingAddressLine2']
#         shippingzipcode = request.form['shippingZipCode']
#         shippingcity = request.form['shippingCity']
#         shippingstate = request.form['shippingState']
#         shippingcountry = request.form['shippingCountry']
#         shippingcontactPerson = request.form['shippingContactPerson']
#         shippingphone = request.form['shippingPhone']
#
#
#
#
#         con = sqlite3.connect('database.db')
#         cur = con.cursor()
#         cur.execute('SELECT email FROM users where email=?', (email,))
#         data = cur.fetchone()
#         con.close()
#
#         if data is None:
#             with sqlite3.connect('database.db') as con:
#                 try:
#                     cur = con.cursor()
#                     # query has been modified here to replace First Name/Last Name
#                     cur.execute(
#                         'INSERT INTO users (password, email, firstName, lastName, address1, address2, zipcode, city, state, country, phone) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
#                         (
#                         hashlib.md5(password.encode()).hexdigest(), email, companyName, "", address1, address2, zipcode,
#                         city, state, country, phone))
#
#                     con.commit()
#
#                     msg = "New Customer Registered Successfully."
#                     return render_template("login.html", error=msg)
#                 except:
#                     con.rollback()
#                     msg = "Error occured during new customer registration!"
#         else:
#             msg = "Emaild already exist in our records, please select another emaildID!"
#         con.close()
#         return render_template("register.html", error=msg)
#

@app.route("/registerationForm")
def registrationForm():
    # return render_template("register.html")
     return  render_template("Reg-6.html")
def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS
def parse(data):
    ans = []
    i = 0
    while i < len(data):
        curr = []
        for j in range(7):
            if i >= len(data):
                break
            curr.append(data[i])
            i += 1
        ans.append(curr)
    return ans
def generate_token(email):
    serializer = URLSafeTimedSerializer(app.config["SECRET_KEY"])
    return serializer.dumps(email, salt=app.config["SECURITY_PASSWORD_SALT"])
def confirm_token(token, expiration=3600):
    serializer = URLSafeTimedSerializer(app.config["SECRET_KEY"])
    try:
        email = serializer.loads(
            token, salt=app.config["SECURITY_PASSWORD_SALT"], max_age=expiration
        )
        return email
    except Exception:
        return False
def send_email(to, subject, template, attachment=None):
    mail = Mail(app)
    msg = Message(
        subject,
        recipients=[to],
        html=template,
        sender=app.config["MAIL_USERNAME"],
    )

    if attachment is not None:
        # Attach a file (example: 'attachment.txt')
        with app.open_resource(attachment) as fp:
            msg.attach(attachment, "text/plain", fp.read())

    mail.send(msg)

# if __name__ == '__main__':
#     # app.run(debug=True)
#     app.run(host="0.0.0.0", port=5000)

